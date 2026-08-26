"""
Celery tasks for Max Pain batch processing and data storage.
"""
import json
import logging
import os
import subprocess
import tempfile
from datetime import datetime, date
from pathlib import Path
from uuid import uuid4
from zoneinfo import ZoneInfo

from sqlalchemy.orm import Session

from ..celery_app import celery_app as app
from ..database import SessionLocal
from ..models.max_pain import MaxPainSnapshot, MaxPainBatch
from ..models.stock_universe import StockUniverse
from ..services.options_snapshot_upsert import upsert_snapshot

logger = logging.getLogger(__name__)

_ET = ZoneInfo("America/New_York")


def _trading_date_for(fetched_at_utc: datetime) -> date:
    """US/Eastern trading day for a naive-UTC fetched_at timestamp.

    See migration 20260805_0028 -- avoids bucketing history queries on a raw
    UTC date_trunc, which drifts across DST/midnight boundaries.
    """
    return fetched_at_utc.replace(tzinfo=ZoneInfo("UTC")).astimezone(_ET).date()


def _chunked(items: list, chunk_size: int):
    """Yield fixed-size chunks from a list."""
    if chunk_size <= 0:
        chunk_size = 200
    for i in range(0, len(items), chunk_size):
        yield items[i:i + chunk_size]


def _resolve_all_active_tickers(db: Session) -> list[dict]:
    """Return all active symbols (with company names) from stock_universe.

    Including the name here lets max_pain_batch.py skip yfinance's `stock.info`
    call per ticker -- by far the slowest yfinance endpoint -- since it's only
    used for a display label that's already in our own database.
    """
    rows = (
        db.query(StockUniverse.symbol, StockUniverse.name)
        .filter(StockUniverse.active_filter())
        .order_by(StockUniverse.symbol.asc())
        .all()
    )
    return [{"symbol": symbol, "company_name": name} for symbol, name in rows if symbol]


def _build_temp_universe_config(
    *,
    tickers: list[dict],
    strike_range_pct: float = 20.0,
    max_strikes: int = 30,
) -> str:
    """Create a temporary max pain config JSON for full-universe runs."""
    payload = {
        "tickers": tickers,
        "strike_range_pct": strike_range_pct,
        "max_strikes": max_strikes,
    }
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False, encoding="utf-8") as tmp:
        json.dump(payload, tmp)
        return tmp.name


def parse_max_pain_output(xlsx_path: str, batch_id: str, db: Session) -> dict:
    """
    Parse Excel output from max_pain_batch.py and store in database.

    NOTE: The Dashboard sheet contains formulas for max pain / totals / ratios.
    Formula cells are often not pre-calculated in generated workbooks, so we
    derive metrics from each ticker detail sheet when dashboard values are empty.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed, cannot parse Excel output")
        return {}

    def _to_float(value):
        if value is None or value == "":
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace('%', '').strip())
        except (TypeError, ValueError):
            return None

    def _to_int(value):
        if value is None or value == "":
            return None
        if isinstance(value, int):
            return value
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None

    def _compute_from_detail_sheet(workbook, ticker: str):
        sheet_name = ticker[:31]
        if sheet_name not in workbook.sheetnames:
            return None
        ws_detail = workbook[sheet_name]
        rows = []
        for detail_row in ws_detail.iter_rows(min_row=5, values_only=True):
            strike, call_oi, put_oi = detail_row[:3]
            strike_f = _to_float(strike)
            call_i = _to_int(call_oi)
            put_i = _to_int(put_oi)
            if strike_f is None:
                continue
            rows.append((strike_f, call_i or 0, put_i or 0))

        if not rows:
            return None

        call_total = sum(r[1] for r in rows)
        put_total = sum(r[2] for r in rows)
        ratio = (put_total / call_total) if call_total else None

        best_strike = None
        best_pain = None
        for strike, _, _ in rows:
            total = 0.0
            for k, c_oi, p_oi in rows:
                total += c_oi * max(strike - k, 0)
                total += p_oi * max(k - strike, 0)
            if best_pain is None or total < best_pain:
                best_pain = total
                best_strike = strike

        return {
            'call_oi': call_total,
            'put_oi': put_total,
            'put_call_ratio': ratio,
            'max_pain': best_strike,
        }

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if 'Dashboard' not in wb.sheetnames:
        logger.error("Dashboard sheet not found in output workbook")
        return {}

    ws = wb['Dashboard']

    tickers_ok = 0
    tickers_failed = 0
    put_call_ratios = []
    closest_strike_dist = None
    closest_ticker = None
    rows_data = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True), start=6):
        if not row or not row[0]:
            continue

        ticker, company_name, status, last_price, expiration, max_pain, distance_pct, call_oi, put_oi, put_call_ratio, note = row[:11]
        if not ticker:
            continue

        try:
            price = _to_float(last_price)
            max_pain_val = _to_float(max_pain)
            distance = _to_float(distance_pct)
            call_oi_val = _to_int(call_oi)
            put_oi_val = _to_int(put_oi)
            ratio = _to_float(put_call_ratio)

            if status == "OK" and (
                max_pain_val is None
                or call_oi_val is None
                or put_oi_val is None
                or ratio is None
            ):
                derived = _compute_from_detail_sheet(wb, str(ticker))
                if derived:
                    if max_pain_val is None:
                        max_pain_val = derived['max_pain']
                    if call_oi_val is None:
                        call_oi_val = derived['call_oi']
                    if put_oi_val is None:
                        put_oi_val = derived['put_oi']
                    if ratio is None:
                        ratio = derived['put_call_ratio']

            if distance is None and price is not None and max_pain_val not in (None, 0):
                distance = ((price - max_pain_val) / max_pain_val) * 100.0

            if status == "OK":
                tickers_ok += 1
                if ratio is not None:
                    put_call_ratios.append(ratio)
                if distance is not None:
                    if closest_strike_dist is None or abs(distance) < abs(closest_strike_dist):
                        closest_strike_dist = distance
                        closest_ticker = ticker
            else:
                tickers_failed += 1

            if isinstance(expiration, datetime):
                expiration_date = expiration.date()
            elif isinstance(expiration, date):
                expiration_date = expiration
            elif expiration:
                expiration_date = datetime.strptime(str(expiration), "%Y-%m-%d").date()
            else:
                expiration_date = None

            fetched_at = datetime.utcnow()
            trading_date_val = _trading_date_for(fetched_at)
            upsert_snapshot(
                db,
                MaxPainSnapshot,
                ticker=ticker,
                trading_date=trading_date_val,
                expiration=expiration_date,
                values={
                    "company_name": company_name,
                    "status": status,
                    "error": note if status == "FAILED" else None,
                    "max_pain_strike": max_pain_val,
                    "call_oi": call_oi_val or 0,
                    "put_oi": put_oi_val or 0,
                    "put_call_ratio": ratio,
                    "last_price": price,
                    "distance_pct": distance,
                    "batch_id": batch_id,
                    "fetched_at": fetched_at,
                },
            )

            rows_data.append({
                'symbol': ticker,
                'company_name': company_name,
                'status': status,
                'expiration': str(expiration_date) if expiration_date else None,
                'call_oi': call_oi_val or 0,
                'put_oi': put_oi_val or 0,
                'put_call_ratio': ratio,
                'max_pain': max_pain_val,
                'distance_pct': distance,
            })
        except (ValueError, TypeError) as e:
            logger.warning(f"Error parsing row {row_idx} for ticker {ticker}: {e}")
            continue

    db.commit()

    avg_ratio = sum(put_call_ratios) / len(put_call_ratios) if put_call_ratios else None

    return {
        "tickers_ok": tickers_ok,
        "tickers_failed": tickers_failed,
        "avg_put_call_ratio": avg_ratio,
        "closest_to_max_pain": closest_ticker,
        "rows": rows_data,
    }


@app.task(name='app.tasks.max_pain_tasks.update_max_pain', bind=True, max_retries=1)
def update_max_pain(self, config_path: str | None = None, wait_until: str | None = None, chain_next: bool = True):
    """
    Run max pain batch job and store results in database.

    Args:
        config_path: Path to config JSON (default: scripts/max_pain_config.json)
        wait_until: HH:MM format to wait until before starting (optional)
        chain_next: If True (the default, used by the daily scheduled run),
            triggers the GEX update when this run finishes. Manual/API-triggered
            runs pass False so an operator re-running Max Pain alone doesn't
            unexpectedly cascade into GEX and Options too.

    Returns:
        dict with task status and summary stats
    """
    batch_id = str(uuid4())[:8]
    db = SessionLocal()
    temp_config_path: str | None = None
    output_path: str | None = None
    
    try:
        logger.info(f"[{batch_id}] Starting max pain batch job")
        
        # Create batch record
        batch = MaxPainBatch(
            id=batch_id,
            status="running",
            started_at=datetime.utcnow(),
            celery_task_id=self.request.id,
        )
        db.add(batch)
        db.commit()
        
        # Prepare script arguments
        if config_path is None:
            tickers = _resolve_all_active_tickers(db)
            if not tickers:
                raise RuntimeError("No active symbols found in stock_universe")
            chunk_size = int(os.getenv("MAX_PAIN_UNIVERSE_CHUNK_SIZE", "50"))
            retry_minutes = float(os.getenv("MAX_PAIN_FULL_RETRY_MINUTES", "1.0"))
            retry_interval = float(os.getenv("MAX_PAIN_FULL_RETRY_INTERVAL_SEC", "10.0"))
            fetch_concurrency = int(os.getenv("MAX_PAIN_FETCH_CONCURRENCY", "6"))
            yfinance_rate_limit = float(os.getenv("MAX_PAIN_YFINANCE_RATE_LIMIT", "3.5"))
            logger.info(
                f"[{batch_id}] Running full-universe max pain for {len(tickers)} active symbols "
                f"in chunks of {chunk_size} (retry window {retry_minutes}m, interval {retry_interval}s, "
                f"fetch concurrency {fetch_concurrency}, rate limit {yfinance_rate_limit}/s)"
            )

            total_ok = 0
            total_failed = 0
            ratio_sum = 0.0
            ratio_count = 0
            closest_ticker = None
            closest_abs_dist = None

            for chunk_index, ticker_chunk in enumerate(_chunked(tickers, chunk_size), start=1):
                chunk_temp_config_path: str | None = None
                chunk_output_path: str | None = None
                try:
                    chunk_temp_config_path = _build_temp_universe_config(tickers=ticker_chunk)

                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                        chunk_output_path = tmp.name

                    args = [
                        "python",
                        "scripts/max_pain_batch.py",
                        chunk_temp_config_path,
                        "--output",
                        chunk_output_path,
                        "--max-retry-minutes",
                        str(retry_minutes),
                        "--retry-interval-sec",
                        str(retry_interval),
                        "--fetch-concurrency",
                        str(fetch_concurrency),
                        "--rate-limit-per-sec",
                        str(yfinance_rate_limit),
                    ]

                    if wait_until and chunk_index == 1:
                        args.extend(["--wait-until", wait_until])

                    logger.info(
                        f"[{batch_id}] Running chunk {chunk_index}: {len(ticker_chunk)} tickers"
                    )
                    result = subprocess.run(
                        args,
                        capture_output=True,
                        text=True,
                        timeout=3600,
                        cwd=str(Path(__file__).resolve().parents[2]),
                    )

                    if result.returncode != 0:
                        logger.error(
                            f"[{batch_id}] Chunk {chunk_index} failed (exit={result.returncode}): "
                            f"{result.stderr[:1000]}"
                        )
                        total_failed += len(ticker_chunk)
                    else:
                        summary = parse_max_pain_output(chunk_output_path, batch_id, db)
                        ok = summary.get("tickers_ok", 0)
                        failed = summary.get("tickers_failed", 0)
                        total_ok += ok
                        total_failed += failed

                        ratio = summary.get("avg_put_call_ratio")
                        if ratio is not None and ok > 0:
                            ratio_sum += ratio * ok
                            ratio_count += ok

                        for row in summary.get("rows", []):
                            if row.get("status") != "OK":
                                continue
                            dist = row.get("distance_pct")
                            if dist is None:
                                continue
                            abs_dist = abs(dist)
                            if closest_abs_dist is None or abs_dist < closest_abs_dist:
                                closest_abs_dist = abs_dist
                                closest_ticker = row.get("symbol")

                    batch.tickers_ok = total_ok
                    batch.tickers_failed = total_failed
                    batch.avg_put_call_ratio = (ratio_sum / ratio_count) if ratio_count else None
                    batch.closest_to_max_pain = closest_ticker
                    db.commit()

                    self.update_state(
                        state='PROGRESS',
                        meta={
                            'current': chunk_index * chunk_size,
                            'total': len(tickers),
                            'percent': round(min(chunk_index * chunk_size, len(tickers)) / len(tickers) * 100, 1),
                            'message': f"Processed chunk {chunk_index}: {total_ok} ok, {total_failed} failed",
                        },
                    )

                finally:
                    try:
                        if chunk_output_path:
                            Path(chunk_output_path).unlink(missing_ok=True)
                        if chunk_temp_config_path:
                            Path(chunk_temp_config_path).unlink(missing_ok=True)
                    except Exception:
                        pass

            # A run where every chunk's subprocess failed used to still be
            # marked "completed" (each per-chunk failure only incremented
            # total_failed, never the overall batch.status), so operators saw
            # a green "Completed" status even though zero real data was
            # produced. Reflect that as a failure instead.
            if total_ok == 0 and total_failed > 0:
                batch.status = "failed"
                batch.error_message = (
                    f"All {total_failed} symbols failed across {chunk_index} chunk(s); "
                    "see worker logs for per-chunk errors."
                )
            else:
                batch.status = "completed"
            batch.completed_at = datetime.utcnow()
            db.commit()

            logger.info(
                f"[{batch_id}] Full-universe batch {batch.status}: {batch.tickers_ok} OK, {batch.tickers_failed} failed"
            )

            return {
                "batch_id": batch_id,
                "status": "success" if batch.status == "completed" else "failed",
                "summary": {
                    "tickers_ok": batch.tickers_ok,
                    "tickers_failed": batch.tickers_failed,
                    "avg_put_call_ratio": batch.avg_put_call_ratio,
                    "closest_to_max_pain": batch.closest_to_max_pain,
                },
            }
        
        # Create temp output file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        args = [
            "python",
            "scripts/max_pain_batch.py",
            config_path,
            "--output", output_path,
            "--fetch-concurrency", os.getenv("MAX_PAIN_FETCH_CONCURRENCY", "6"),
            "--rate-limit-per-sec", os.getenv("MAX_PAIN_YFINANCE_RATE_LIMIT", "3.5"),
        ]
        
        if wait_until:
            args.extend(["--wait-until", wait_until])
        
        # Run batch script
        logger.info(f"[{batch_id}] Running: {' '.join(args)}")
        result = subprocess.run(
            args,
            capture_output=True,
            text=True,
            timeout=3600,  # 1 hour timeout
            cwd=str(Path(__file__).resolve().parents[2]),
        )
        
        if result.returncode != 0:
            error_msg = f"Batch job failed: {result.stderr}"
            logger.error(f"[{batch_id}] {error_msg}")
            batch.status = "failed"
            batch.error_message = error_msg
            db.commit()
            return {
                "batch_id": batch_id,
                "status": "failed",
                "error": error_msg,
            }
        
        logger.info(f"[{batch_id}] Batch job completed, parsing results")
        
        # Parse results and store in database
        summary = parse_max_pain_output(output_path, batch_id, db)
        
        # Update batch record with summary
        batch.status = "completed"
        batch.completed_at = datetime.utcnow()
        batch.tickers_ok = summary.get("tickers_ok", 0)
        batch.tickers_failed = summary.get("tickers_failed", 0)
        batch.avg_put_call_ratio = summary.get("avg_put_call_ratio")
        batch.closest_to_max_pain = summary.get("closest_to_max_pain")
        db.commit()
        
        logger.info(
            f"[{batch_id}] Batch completed: "
            f"{batch.tickers_ok} OK, {batch.tickers_failed} failed"
        )
        
        # Cleanup temp file
        if output_path:
            Path(output_path).unlink(missing_ok=True)
        if temp_config_path:
            Path(temp_config_path).unlink(missing_ok=True)
        
        return {
            "batch_id": batch_id,
            "status": "success",
            "summary": summary,
        }
    
    except subprocess.TimeoutExpired:
        logger.error(f"[{batch_id}] Batch job timed out")
        batch = db.query(MaxPainBatch).filter(MaxPainBatch.id == batch_id).first()
        if batch:
            batch.status = "failed"
            batch.error_message = "Timeout after 1 hour"
            db.commit()
        return {"batch_id": batch_id, "status": "timeout"}
    
    except Exception as e:
        logger.exception(f"[{batch_id}] Unexpected error in batch job")
        try:
            batch = db.query(MaxPainBatch).filter(MaxPainBatch.id == batch_id).first()
            if batch:
                batch.status = "failed"
                batch.error_message = str(e)
                db.commit()
        except Exception:
            pass
        return {"batch_id": batch_id, "status": "error", "error": str(e)}
    
    finally:
        try:
            if output_path:
                Path(output_path).unlink(missing_ok=True)
            if temp_config_path:
                Path(temp_config_path).unlink(missing_ok=True)
        except Exception:
            pass
        db.close()

        # Chain into the GEX update regardless of this run's outcome, so a bad
        # Max Pain day doesn't stall the rest of the evening pipeline. Explicit
        # queue routing is kept here even though update_gex now has a
        # celery_app.py task_routes backstop, since that default only saves
        # a caller that forgets to pass queue= -- it's not a reason to
        # start omitting it deliberately.
        if chain_next:
            try:
                from .gex_tasks import update_gex
                from .market_queues import data_fetch_queue_for_market
                update_gex.apply_async(queue=data_fetch_queue_for_market("US"))
                logger.info(f"[{batch_id}] Chained GEX update")
            except Exception:
                logger.exception(f"[{batch_id}] Failed to chain GEX update")


@app.task(name='app.tasks.max_pain_tasks.schedule_daily_update')
def schedule_daily_update():
    """
    Scheduled task to run daily max pain update.

    Add to celery beat schedule:

    from celery.schedules import crontab
    app.conf.beat_schedule = {
        'max-pain-nightly': {
            'task': 'app.tasks.max_pain_tasks.schedule_daily_update',
            'schedule': crontab(hour=22, minute=30),  # 22:30 every day
        },
    }
    """
    logger.info("Triggering scheduled max pain update")
    # Explicit queue routing (not just .delay()), kept even though
    # update_max_pain now has a celery_app.py task_routes backstop -- routing
    # the *wrapper* task (this function) to data_fetch_us -- via the beat
    # entry's 'options': {'queue': ...} or SCHEDULED_TASKS'
    # manual_dispatch_options -- does NOT propagate to update_max_pain: a
    # plain .delay() call from inside a task dispatches on that task's own
    # route, not the caller's queue.
    from .market_queues import data_fetch_queue_for_market
    update_max_pain.apply_async(queue=data_fetch_queue_for_market("US"))
